import asyncio
import os

# dotenv loads the environment variables from the file ".env"
from dotenv import load_dotenv

# Load the Call of Duty Module (this is the folder "callofduty")
# CallofDuty.py is an asynchronous, object-oriented Python wrapper for the Call of Duty API.
import callofduty as cod

# Openpyxl Read & Write xlsx-Files (Excel, Libreoffice calc) with python
import openpyxl


# This is the main function. Program execution starts at the bottom of this file (line 316)
async def main():
    # Load Environment Variables from the file ".env"
    load_dotenv()
    cod_user_email = os.environ["ATVI_EMAIL"]
    cod_user_pass = os.environ["ATVI_PASSWORD"]

    # you may change this to any valid user
    userNameForQueries = "neanderHigh"
    # userNameForQueries = "Icoryx#284"

    # Print all your env variables
    # print(os.environ)

    # Print the PATH env
    # print(os.environ['PATH'])

    print(f"User Email: {cod_user_email}, User Password: {cod_user_pass}")

    # Create Excel
    my_xls = "./cod_modern_warfare.xlsx"
    wb = openpyxl.Workbook()

    # CoD Login:
    client = await cod.Login(cod_user_email, cod_user_pass)

    # Show all Map of Modern Warfare
    wb_sheet_map = wb.create_sheet("Maps")
    maps = await client.GetAvailableMaps(cod.Title.ModernWarfare)
    # Sheet Header
    wb_sheet_map.append(("Map Name", "Mode"))
    for mapName in maps:
        for mode in maps[mapName]:
            print(f"{mapName} - {mode}")
            wb_sheet_map.append((mapName, mode))

    # Seasons & Tiers
    wb_sheet_seasons = wb.create_sheet("Seasons")
    wb_sheet_seasons.append(
        (
            "Season Title",
            "Season Name",
            "Tier",
            "Tier Name",
            "Tier Rarity",
            "Tier Category",
        )
    )
    x = 1
    while True:
        season = await client.GetLootSeason(cod.Title.ModernWarfare, x)
        if not season.name:
            break
        x += 1
        for tier in season.tiers:
            print(
                f"Season: {season.title.name}: {season.name} - Tier {tier.tier}: {tier.name} - {tier.rarity} {tier.category}"
            )
            wb_sheet_seasons.append(
                (
                    season.title.name,
                    season.name,
                    tier.tier,
                    tier.name,
                    tier.rarity,
                    tier.category,
                )
            )
        # for chase in season.chase:
        #     print(f"Chase: {chase.name} - {chase.rarity} {chase.category}")

    # News
    wb_sheet_news = wb.create_sheet("News")
    news = await client.GetNewsFeed(limit=10)
    for post in news:
        print(f"{post.published.date()}: {post.title}")
        wb_sheet_news.append((post.published.date(), post.title))

    # Delete the default sheet
    del wb["Sheet"]

    # Show sheet names
    print(wb.sheetnames)

    # Authentication required
    print("Friends Request")
    requests = await client.GetMyFriendRequests()
    for incoming in requests["incoming"]:
        print(
            f"Incoming Friend Request: {incoming.username} ({incoming.platform.name})"
        )
    for outgoing in requests["outgoing"]:
        print(
            f"Outgoing Friend Request: {outgoing.username} ({outgoing.platform.name})"
        )

    friends = await client.GetMyFriends()
    for friend in friends:
        print(
            f"{friend.platform.name}: {friend.username} ({friend.accountId}), Online? {friend.online}"
        )
        for identity in friend.identities:
            print(
                f" - {identity.platform.name}: {identity.username} ({identity.accountId})"
            )

    identities = await client.GetMyIdentities()
    for identity in identities:
        title = identity["title"].name
        username = identity["username"]
        platform = identity["platform"].name
        print(f"{title}: {username} ({platform})")

    # accounts = await client.GetMyAccounts()
    # for account in accounts:
    #     print(f"{account.username} ({account.platform.name})")

    # player = await client.GetPlayer(cod.Platform.BattleNet, "Mxtive#1930")
    # print(f"{player.username} ({player.platform.name})")

    # player = await client.GetPlayer(cod.Platform.BattleNet, "Mxtive#1930")
    # summary = await player.matchesSummary(cod.Title.ModernWarfare, cod.Mode.Warzone, limit=20)
    # print(summary)

    print("Videos")
    videos = await client.GetVideoFeed(limit=3)
    for video in videos:
        print(f"{video.title} - {video.url}")

    print("Leaderboard")
    leaderboard = await client.GetLeaderboard(
        cod.Title.ModernWarfare, cod.Platform.BattleNet, gameMode="cyber", page=1
    )
    for entry in leaderboard.entries:
        print(f"#{entry.rank}: {entry.username} ({entry.platform.name})")

    leaderboard = await client.GetLeaderboard(
        cod.Title.ModernWarfare, cod.Platform.BattleNet, gameMode="cyber", page=2
    )
    for entry in leaderboard.entries:
        print(f"#{entry.rank}: {entry.username} ({entry.platform.name})")

    # Not allowed
    # leaderboard = await client.GetPlayerLeaderboard(
    #     cod.Title.BlackOps4, cod.Platform.Steam, "neanderHigh"
    # )
    # for entry in leaderboard.entries:
    #     if entry.username == "neanderHigh":
    #         print(f"#{entry.rank}: {entry.username} ({entry.platform.name})")

    print("Leaderboard WW2")
    player = await client.GetPlayer(cod.Platform.Steam, userNameForQueries)
    try:
        leaderboard = await player.leaderboard(cod.Title.WWII)
        for entry in leaderboard.entries:
            if entry.username == player.username:
                print(f"#{entry.rank}: {entry.username} ({entry.platform.name})")
    except cod.errors.HTTPException:
        print("No entry for this user!")

    print("Next try")
    feed = await client.GetFriendFeed(limit=3)
    for item in feed:
        print(f"[{item.date.strftime('%Y-%m-%d %H:%M')}] {item.text}")
        if (match := item.match) is not None:
            for team in await match.teams():
                for player in team:
                    if player.username != item.player.username:
                        print(f"       {player.username} ({player.platform.name})")

    # Feed
    print("Feed readct")
    feed = await client.GetFriendFeed(limit=5)
    for item in feed:
        print(item.text)
        await item.react(cod.Reaction.Fire)

    print("Feed unreadct")
    feed = await client.GetFriendFeed(limit=5)
    for item in feed:
        print(item.text)
        await item.unreact()

    print("Feed item favorite")
    feed = await client.GetFriendFeed(limit=1)
    for item in feed:
        print(item.text)
        test = await item.favorite()
        print(test)

    print("Feed item unfavorite")
    feed = await client.GetFriendFeed(limit=1)
    for item in feed:
        print(item.text)
        await item.unfavorite()

    print("Matches")
    matches = await client.GetPlayerMatches(
        cod.Platform.Activision,
        "Yeah#8649242",
        cod.Title.ModernWarfare,
        cod.Mode.Multiplayer,
        limit=3,
    )
    for match in matches:
        print(match.id, match.title, match.type)

    # player = await client.GetPlayer(cod.Platform.BattleNet, "Mxtive#1930")
    # match = (await player.matches(cod.Title.ModernWarfare, cod.Mode.Multiplayer, limit=3))[1]
    # match = await client.GetMatch(cod.Title.ModernWarfare, cod.Platform.Activision, match.id)
    # teams = await match.teams()
    # for team in teams:
    #     for player in team:
    #         print(player.username)
    # details = await match.details()
    # print(details)

    # results = await client.SearchPlayers(cod.Platform.Activision, "Tustin")
    # for player in results:
    #     print(f"{player.username} ({player.platform.name})")

    # player = await client.GetPlayer(cod.Platform.BattleNet, "Mxtive#1930")
    # profile = await player.profile(cod.Title.ModernWarfare, cod.Mode.Multiplayer)
    # print(profile)

    # localize = await client.GetLocalize()
    # print(localize)

    # Does not work with neanderHigh (KeyError: 'loadouts')
    print("Loadouts")
    loadouts = await client.GetPlayerLoadouts(
        cod.Platform.Steam, userNameForQueries, cod.Title.BlackOps4
    )
    for loadout in loadouts:
        if loadout.name != "":
            print(f"Class: {loadout.name} (Unlocked: {loadout.unlocked})")
            if loadout.primary.id is not None:
                print(
                    f" - Primary Weapon: {loadout.primary.id} (Variant: {loadout.primary.variant})"
                )
                print(f"  - Camo: {loadout.primary.camo}")
                for attachment in loadout.primary.attachments:
                    if attachment.id is not None:
                        print(f"  - Attachment: {attachment.id}")
            if loadout.secondary.id is not None:
                print(
                    f" - Secondary Weapon: {loadout.secondary.id} (Variant: {loadout.secondary.variant})"
                )
                print(f"  - Camo: {loadout.secondary.camo}")
                for attachment in loadout.secondary.attachments:
                    if attachment.id is not None:
                        print(f"  - Attachment: {attachment.id}")
            for equipment in loadout.equipment:
                if equipment.id is not None:
                    print(f" - Equipment: {equipment.id}")
            for perk in loadout.perks:
                if perk.id is not None:
                    print(f" - Perk: {perk.id}")
            for wildcard in loadout.wildcards:
                if wildcard.id is not None:
                    print(f" - Wildcard: {wildcard.id}")

    # player = await client.GetPlayer(cod.Platform.PlayStation, "ImMotive__")
    # loadouts = await player.loadouts(cod.Title.BlackOps4)
    # for loadout in loadouts:
    #     if loadout.name != "":
    #         print(f"Class: {loadout.name} (Unlocked: {loadout.unlocked})")

    # unlocks = await client.GetPlayerLoadoutUnlocks(cod.Platform.PlayStation, "ImMotive__", cod.Title.BlackOps4)
    # for unlock in unlocks:
    #     print(unlock.id)

    # player = await client.GetPlayer(cod.Platform.PlayStation, "ImMotive__")
    # unlocks = await player.loadoutUnlocks(cod.Title.BlackOps4)
    # for unlock in unlocks:
    #     print(unlock.id)

    # stamp = await client.GetAuthenticityStamp(
    #     cod.Platform.BattleNet, "Slicky#21337", "Swiftly Snarling Gamy Generators"
    # )
    # print(stamp.data)

    # player = await client.GetPlayer(cod.Platform.BattleNet, "Slicky#21337")
    # stamp = await player.authenticityStamp("Swiftly Snarling Gamy Generators")
    # print(stamp.stats)

    # req = await client.AddFriend(5273496286943517033)
    # print(f"Friend Request Status: {req}")

    # req = await client.RemoveFriend(13940176918450289589)
    # print(f"Friend Request Status: {req}")

    # results = await client.SearchPlayers(cod.Platform.Activision, "Tustin")
    # for player in results:
    #     print(f"{player.username} ({player.platform.name})")
    #     if player.username == "Tustin#1365515":
    #         req = await player.removeFriend()
    #         print(f"Removed Friend ({req})")
    #         req = await player.addFriend()
    #         print(f"Added Friend ({req})")

    # favs = await client.GetMyFavorites()
    # for favorite in favs:
    #     print(f"Favorite: {favorite.username} ({favorite.platform.name})")

    # favs = await client.AddFavorite(cod.Platform.Activision, "Dad#1869899")
    # print(f"Favorites: {len(favs)}")

    # player = await client.GetPlayer(cod.Platform.Activision, "Dad#1869899")
    # favs = await player.removeFavorite()
    # print(f"Favorites: {len(favs)}")

    # results = await client.SearchPlayers(cod.Platform.Activision, "Tustin")
    # for player in results:
    #     if player.username == "Tustin#1365515":
    #         await player.block()
    #         await player.unblock()
    #         req = await player.addFriend()
    #         print(req)

    # squad = await client.GetSquad("Autists")
    # print(f"{squad.name} - {squad.description}")
    # print(f"Owner: {squad.owner.username} ({squad.owner.platform.name})")
    # for member in squad.members:
    #     if member.username != squad.owner.username:
    #         print(f"Member: {member.username} ({member.platform.name})")

    # squad = await client.GetPlayerSquad(cod.Platform.Activision, "Yeah#8649242")
    # print(f"{squad.name} - {squad.description}")
    # print(f"Owner: {squad.owner.username} ({squad.owner.platform.name})")
    # for member in squad.members:
    #     if member.username != squad.owner.username:
    #         print(f"Member: {member.username} ({member.platform.name})")

    # squad = await client.GetMySquad()
    # print(f"{squad.name} - {squad.description}")
    # print(f"Owner: {squad.owner.username} ({squad.owner.platform.name})")
    # for member in squad.members:
    #     if member.username != squad.owner.username:
    #         print(f"Member: {member.username} ({member.platform.name})")

    # print(f"Leaving Squad '{squad.name}''...")
    # squad = await client.LeaveSquad()
    # print(f"Current Squad: {squad.name} - {squad.description} (Members: {len(squad.members)})")

    # squad = await client.GetSquad("Hmmmm")
    # print(f"Joining Squad '{squad.name}'...")
    # await squad.join()
    # squad = await client.GetMySquad()
    # print(f"Current Squad: {squad.name} - {squad.description} (Members: {len(squad.members)})")

    # squad = await client.GetSquad("Hmmmm")
    # await squad.report()

    # challenge = await client.GetSquadsTournament(cod.Title.ModernWarfare)
    # print(f"{challenge.title.name} Squads Tournament: {challenge.name} - {challenge.description}")

    # Save the xlsx (workbook)
    wb.save(my_xls)


# Program Start:
asyncio.get_event_loop().run_until_complete(main())
